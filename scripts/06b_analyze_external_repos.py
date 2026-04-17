"""File-listing + README extraction for non-openICPSR repos.

Handles the deposits landed by scripts 03b / 03c / 04b:
  - dataverse (10.7910/DVN/... and other Dataverse installations)
  - zenodo    (10.5281/zenodo.NNN)
  - mendeley  (10.17632/XXX.V or https://data.mendeley.com/datasets/XXX/V)

For each repo: fetches file metadata via JSON API, populates repo_files
(with repo_host), finds a README, downloads and extracts its text, and
populates readme_analysis.
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
import time
from pathlib import Path
from typing import Any
from urllib.parse import quote

import requests

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

try:
    from config import (
        DATA_DIR,
        FILE_TYPE_CLASSIFICATIONS,
        OPENALEX_EMAIL,
        RAW_DIR,
        RESTRICTION_INDICATORS,
    )
    from db import get_connection, init_db
except ImportError:  # pragma: no cover
    from scripts.config import (
        DATA_DIR,
        FILE_TYPE_CLASSIFICATIONS,
        OPENALEX_EMAIL,
        RAW_DIR,
        RESTRICTION_INDICATORS,
    )
    from scripts.db import get_connection, init_db

LOGGER = logging.getLogger(__name__)

RATE_LIMIT_SLEEP = 0.5
REQUEST_TIMEOUT = 30
LARGE_TXT_THRESHOLD = 100 * 1024
README_RE = re.compile(r"read[\s_-]?me", re.IGNORECASE)

EXTERNAL_CACHE_DIR = RAW_DIR / "external_repos"


def configure_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )


def build_session() -> requests.Session:
    session = requests.Session()
    contact = OPENALEX_EMAIL or "unknown@example.com"
    session.headers.update({
        "Accept": "application/json",
        "User-Agent": f"ReplicationTracker/1.0 (mailto:{contact})",
    })
    return session


# ---------------------------------------------------------------------------
# Identifier normalization
# ---------------------------------------------------------------------------

DATAVERSE_DOI_RE = re.compile(r"10\.7910/dvn/([a-z0-9]+)", re.IGNORECASE)
DATAVERSE_URL_RE = re.compile(
    r"dataverse\.harvard\.edu/(?:dataset\.xhtml\?persistentId=)?doi:([^\s&#]+)",
    re.IGNORECASE,
)
ZENODO_DOI_RE = re.compile(r"10\.5281/zenodo\.(\d+)", re.IGNORECASE)
ZENODO_URL_RE = re.compile(r"zenodo\.org/records?/(\d+)", re.IGNORECASE)
MENDELEY_DOI_RE = re.compile(r"10\.17632/([a-z0-9]+)(?:\.\d+)?", re.IGNORECASE)
MENDELEY_URL_RE = re.compile(
    r"data\.mendeley\.com/datasets/([a-z0-9]+)(?:/\d+)?",
    re.IGNORECASE,
)


def dataverse_pid(repo_doi: str) -> str | None:
    """Return a Dataverse persistentId string `doi:10.7910/DVN/XXX` or None."""
    low = repo_doi.lower()
    m = DATAVERSE_URL_RE.search(low)
    if m:
        return f"doi:{m.group(1).upper()}"
    m = DATAVERSE_DOI_RE.search(low)
    if m:
        return f"doi:10.7910/DVN/{m.group(1).upper()}"
    # Other Dataverse installations — keep DOI as-is (will fail for non-Harvard though)
    if low.startswith("10.") and "/" in low:
        return f"doi:{repo_doi}"
    return None


def zenodo_record_id(repo_doi: str) -> str | None:
    m = ZENODO_DOI_RE.search(repo_doi.lower())
    if m:
        return m.group(1)
    m = ZENODO_URL_RE.search(repo_doi.lower())
    if m:
        return m.group(1)
    return None


def mendeley_dataset_id(repo_doi: str) -> str | None:
    m = MENDELEY_DOI_RE.search(repo_doi.lower())
    if m:
        return m.group(1)
    m = MENDELEY_URL_RE.search(repo_doi.lower())
    if m:
        return m.group(1)
    return None


# ---------------------------------------------------------------------------
# File classification (shared with 06_analyze_repos.py)
# ---------------------------------------------------------------------------

def classify_file(filename: str, size_bytes: int) -> dict[str, Any]:
    ext = Path(filename).suffix.lower()
    is_readme = bool(README_RE.search(filename) and not filename.startswith("._"))
    if is_readme:
        file_type = "documentation"
    elif ext == ".txt" and size_bytes > LARGE_TXT_THRESHOLD:
        file_type = "data"
    else:
        file_type = FILE_TYPE_CLASSIFICATIONS.get(ext, "other")
    return {
        "filename": filename,
        "extension": ext,
        "file_type": file_type,
        "size_bytes": size_bytes,
        "is_readme": is_readme,
    }


def extract_readme_text(raw_bytes: bytes, filename: str) -> str | None:
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        try:
            import fitz
            pages = []
            with fitz.open(stream=raw_bytes, filetype="pdf") as doc:
                for p in doc:
                    t = p.get_text()
                    if t:
                        pages.append(t)
            return "\n\n".join(pages) if pages else None
        except Exception as exc:
            LOGGER.warning("PDF extract failed (%s): %s", filename, exc)
            return None
    if ext == ".docx":
        try:
            import fitz
            pages = []
            with fitz.open(stream=raw_bytes, filetype="docx") as doc:
                for p in doc:
                    t = p.get_text()
                    if t:
                        pages.append(t)
            return "\n\n".join(pages) if pages else None
        except Exception as exc:
            LOGGER.warning("DOCX extract failed (%s): %s", filename, exc)
            return None
    if ext in (".txt", ".md", ".rst", ""):
        for enc in ("utf-8", "latin-1"):
            try:
                return raw_bytes.decode(enc)
            except UnicodeDecodeError:
                continue
        return None
    if ext == ".rtf":
        try:
            from striprtf.striprtf import rtf_to_text
            for enc in ("utf-8", "latin-1"):
                try:
                    raw_str = raw_bytes.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                return None
            text = rtf_to_text(raw_str, errors="ignore")
            return text or None
        except Exception as exc:
            LOGGER.warning("RTF extract failed (%s): %s", filename, exc)
            return None
    if ext in (".html", ".htm"):
        try:
            from bs4 import BeautifulSoup
            for enc in ("utf-8", "latin-1"):
                try:
                    raw_str = raw_bytes.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                return None
            soup = BeautifulSoup(raw_str, "html.parser")
            text = soup.get_text("\n", strip=True)
            return text or None
        except Exception as exc:
            LOGGER.warning("HTML extract failed (%s): %s", filename, exc)
            return None
    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
            from io import BytesIO
            wb = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
            parts: list[str] = []
            for sheet in wb.worksheets:
                parts.append(f"[{sheet.title}]")
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        parts.append("\t".join(cells))
            text = "\n".join(parts)
            return text or None
        except Exception as exc:
            LOGGER.warning("XLSX extract failed (%s): %s", filename, exc)
            return None
    LOGGER.warning("Unsupported README ext for %s: %s", filename, ext)
    return None


# ---------------------------------------------------------------------------
# Host adapters
# ---------------------------------------------------------------------------


def fetch_dataverse(session: requests.Session, repo_doi: str) -> dict[str, Any] | None:
    pid = dataverse_pid(repo_doi)
    if not pid:
        return None
    url = "https://dataverse.harvard.edu/api/datasets/:persistentId/versions/:latest/files"
    resp = session.get(url, params={"persistentId": pid}, timeout=REQUEST_TIMEOUT)
    if resp.status_code != 200:
        return None
    payload = resp.json()
    if payload.get("status") != "OK":
        return None

    files: list[dict[str, Any]] = []
    readme_download: tuple[str, str] | None = None  # (filename, url)
    for entry in payload.get("data", []):
        df = entry.get("dataFile") or {}
        name = df.get("filename")
        size = df.get("filesize") or 0
        if not name:
            continue
        files.append({"filename": name, "size_bytes": int(size)})
        if readme_download is None and README_RE.search(name) and not name.startswith("._"):
            fid = df.get("id")
            if fid:
                readme_download = (
                    name,
                    f"https://dataverse.harvard.edu/api/access/datafile/{fid}",
                )
    return {"files": files, "readme": readme_download}


def fetch_zenodo(session: requests.Session, repo_doi: str) -> dict[str, Any] | None:
    rid = zenodo_record_id(repo_doi)
    if not rid:
        return None
    url = f"https://zenodo.org/api/records/{rid}"
    resp = session.get(url, timeout=REQUEST_TIMEOUT)
    if resp.status_code != 200:
        return None
    payload = resp.json()
    files: list[dict[str, Any]] = []
    readme_download: tuple[str, str] | None = None
    for f in payload.get("files") or []:
        name = f.get("key")
        size = f.get("size") or 0
        if not name:
            continue
        files.append({"filename": name, "size_bytes": int(size)})
        if readme_download is None and README_RE.search(name) and not name.startswith("._"):
            dl = (f.get("links") or {}).get("self")
            if dl:
                readme_download = (name, dl)
    return {"files": files, "readme": readme_download}


def fetch_mendeley(session: requests.Session, repo_doi: str) -> dict[str, Any] | None:
    did = mendeley_dataset_id(repo_doi)
    if not did:
        return None
    url = f"https://data.mendeley.com/public-api/datasets/{did}"
    resp = session.get(url, timeout=REQUEST_TIMEOUT)
    if resp.status_code != 200:
        return None
    payload = resp.json()
    files: list[dict[str, Any]] = []
    readme_download: tuple[str, str] | None = None
    for f in payload.get("files") or []:
        name = f.get("filename")
        size = f.get("size") or (f.get("content_details") or {}).get("size") or 0
        if not name:
            continue
        files.append({"filename": name, "size_bytes": int(size)})
        if readme_download is None and README_RE.search(name) and not name.startswith("._"):
            dl = (f.get("content_details") or {}).get("download_url")
            if dl:
                readme_download = (name, dl)
    return {"files": files, "readme": readme_download}


HOST_ADAPTERS = {
    "dataverse": fetch_dataverse,
    "zenodo": fetch_zenodo,
    "mendeley": fetch_mendeley,
}


# ---------------------------------------------------------------------------
# README fetch + caching
# ---------------------------------------------------------------------------

def safe_cache_name(repo_doi: str, filename: str) -> str:
    safe = re.sub(r"[^a-zA-Z0-9._-]+", "_", repo_doi) + "__" + re.sub(
        r"[^a-zA-Z0-9._-]+", "_", filename
    )
    return safe[:180]


def fetch_readme(session: requests.Session, repo_doi: str,
                 host: str, filename: str, url: str,
                 cache_dir: Path) -> bytes | None:
    cache_dir.mkdir(parents=True, exist_ok=True)
    path = cache_dir / safe_cache_name(repo_doi, filename)
    if path.exists():
        return path.read_bytes()
    try:
        resp = session.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        if resp.status_code != 200:
            LOGGER.warning("README fetch %s -> HTTP %d", url, resp.status_code)
            return None
        path.write_bytes(resp.content)
        return resp.content
    except requests.RequestException as exc:
        LOGGER.warning("README fetch failed for %s: %s", url, exc)
        return None


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def get_pending_repos(conn, hosts: list[str]) -> list[dict[str, str]]:
    placeholders = ",".join("?" * len(hosts))
    rows = conn.execute(
        f"""
        SELECT DISTINCT repo_doi, repo_host
        FROM repo_mappings
        WHERE repo_host IN ({placeholders})
          AND repo_doi NOT IN (SELECT DISTINCT repo_doi FROM repo_files WHERE repo_doi IS NOT NULL)
        ORDER BY repo_host, repo_doi
        """,
        hosts,
    ).fetchall()
    return [{"repo_doi": r["repo_doi"], "repo_host": r["repo_host"]} for r in rows]


def get_readme_retry_repos(conn, hosts: list[str]) -> list[dict[str, str]]:
    """Repos that have a README on disk but no extracted text — candidates for retry."""
    placeholders = ",".join("?" * len(hosts))
    rows = conn.execute(
        f"""
        SELECT DISTINCT rm.repo_doi, rm.repo_host
        FROM repo_mappings rm
        JOIN readme_analysis ra ON ra.repo_doi = rm.repo_doi
        WHERE rm.repo_host IN ({placeholders})
          AND ra.has_readme = 1
          AND (ra.readme_text IS NULL OR ra.readme_text = '')
        ORDER BY rm.repo_host, rm.repo_doi
        """,
        hosts,
    ).fetchall()
    return [{"repo_doi": r["repo_doi"], "repo_host": r["repo_host"]} for r in rows]


def insert_files(conn, repo_doi: str, host: str,
                 classified: list[dict[str, Any]]) -> None:
    with conn:
        for f in classified:
            conn.execute(
                """
                INSERT INTO repo_files
                    (repo_doi, repo_host, filename, extension, file_type, size_bytes)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (repo_doi, host, f["filename"], f["extension"], f["file_type"], f["size_bytes"]),
            )


def insert_readme(conn, repo_doi: str, host: str,
                  filename: str | None, text: str | None) -> None:
    has_readme = filename is not None
    flags: list[str] = []
    if text:
        low = text.lower()
        for phrase in RESTRICTION_INDICATORS:
            if phrase.lower() in low:
                flags.append(phrase)
    with conn:
        conn.execute(
            """
            INSERT OR REPLACE INTO readme_analysis
                (repo_doi, repo_host, has_readme, readme_text, restriction_flags, restriction_count)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (repo_doi, host, has_readme, text, json.dumps(flags), len(flags)),
        )


def insert_unavailable(conn, repo_doi: str, host: str) -> None:
    with conn:
        conn.execute(
            """
            INSERT INTO repo_files (repo_doi, repo_host, filename, extension, file_type, size_bytes)
            VALUES (?, ?, '__unavailable__', NULL, 'unavailable', NULL)
            """,
            (repo_doi, host),
        )
    insert_readme(conn, repo_doi, host, None, None)


def run_readme_retry(conn, repos: list[dict[str, str]]) -> int:
    session = build_session()
    retried = 0
    new_extracts = 0
    for idx, rep in enumerate(repos, 1):
        repo_doi = rep["repo_doi"]
        host = rep["repo_host"]
        if idx > 1:
            time.sleep(RATE_LIMIT_SLEEP)
        try:
            listing = HOST_ADAPTERS[host](session, repo_doi)
        except requests.RequestException as exc:
            LOGGER.warning("%s listing failed for %s: %s", host, repo_doi, exc)
            continue
        if not listing:
            continue
        readme_info = listing.get("readme")
        if not readme_info:
            continue
        name, url = readme_info
        raw = fetch_readme(session, repo_doi, host, name, url,
                           EXTERNAL_CACHE_DIR / host)
        text = extract_readme_text(raw, name) if raw else None
        if text:
            new_extracts += 1
        insert_readme(conn, repo_doi, host, name, text)
        retried += 1
        if idx % 10 == 0:
            LOGGER.info("  retry progress: %d/%d new_extracts=%d",
                        idx, len(repos), new_extracts)

    LOGGER.info("Retried: %d | new text extracted: %d", retried, new_extracts)
    return 0


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    configure_logging()
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--hosts",
        default="dataverse,zenodo,mendeley",
        help="Comma-separated host labels to process",
    )
    parser.add_argument("--limit", type=int, default=None,
                        help="Max repos to process")
    parser.add_argument("--retry-readmes", action="store_true",
                        help="Skip file listing; only re-fetch/extract missing README text")
    args = parser.parse_args()

    hosts = [h.strip() for h in args.hosts.split(",") if h.strip()]
    unknown = [h for h in hosts if h not in HOST_ADAPTERS]
    if unknown:
        LOGGER.error("No adapter for hosts: %s", unknown)
        return 1

    EXTERNAL_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    db_path = init_db()
    conn = get_connection(db_path)

    try:
        if args.retry_readmes:
            repos = get_readme_retry_repos(conn, hosts)
            if args.limit:
                repos = repos[:args.limit]
            LOGGER.info("README-retry candidates: %d (hosts=%s)", len(repos), hosts)
            return run_readme_retry(conn, repos)

        repos = get_pending_repos(conn, hosts)
        if args.limit:
            repos = repos[:args.limit]
        LOGGER.info("Pending repos: %d (hosts=%s)", len(repos), hosts)

        ok = fail = files_total = readmes_found = readmes_extracted = 0
        per_host: dict[str, dict[str, int]] = {h: {"ok": 0, "fail": 0} for h in hosts}
        session = build_session()

        for idx, rep in enumerate(repos, 1):
            repo_doi = rep["repo_doi"]
            host = rep["repo_host"]
            if idx > 1:
                time.sleep(RATE_LIMIT_SLEEP)

            try:
                listing = HOST_ADAPTERS[host](session, repo_doi)
            except requests.RequestException as exc:
                LOGGER.warning("%s fetch failed for %s: %s", host, repo_doi, exc)
                listing = None

            if listing is None or not listing.get("files"):
                fail += 1
                per_host[host]["fail"] += 1
                insert_unavailable(conn, repo_doi, host)
                continue

            classified = [classify_file(f["filename"], f["size_bytes"])
                          for f in listing["files"]]
            insert_files(conn, repo_doi, host, classified)
            files_total += len(classified)
            ok += 1
            per_host[host]["ok"] += 1

            readme_info = listing.get("readme")
            readme_text: str | None = None
            readme_name: str | None = None
            if readme_info:
                readmes_found += 1
                readme_name, readme_url = readme_info
                raw = fetch_readme(session, repo_doi, host, readme_name,
                                   readme_url, EXTERNAL_CACHE_DIR / host)
                if raw:
                    readme_text = extract_readme_text(raw, readme_name)
                    if readme_text:
                        readmes_extracted += 1

            insert_readme(conn, repo_doi, host, readme_name, readme_text)

            if idx % 25 == 0:
                LOGGER.info(
                    "Progress: %d/%d ok=%d fail=%d files=%d readmes=%d/%d",
                    idx, len(repos), ok, fail, files_total,
                    readmes_extracted, readmes_found,
                )

        LOGGER.info("=" * 60)
        LOGGER.info("Pending: %d | ok=%d fail=%d", len(repos), ok, fail)
        LOGGER.info("Files inserted: %d", files_total)
        LOGGER.info("READMEs: found=%d extracted=%d", readmes_found, readmes_extracted)
        LOGGER.info("Per host:")
        for h, stats in per_host.items():
            LOGGER.info("  %-10s ok=%d fail=%d", h, stats["ok"], stats["fail"])
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
